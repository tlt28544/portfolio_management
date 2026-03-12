#!/usr/bin/env python3
import json
import os
import re
import shutil
import tempfile
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple

from google.cloud import storage
from google.oauth2 import service_account
from googleapiclient.discovery import build
from openpyxl import load_workbook

PORTFOLIO_DEFAULT = "759668"
FUND_DEFAULT = "759668S"
TRADE_STATUS_DEFAULT = "New"
INVESTMENT_TYPE_DEFAULT = "Equity"
BROKER_CODE_DEFAULT = "CMSHK"
CUSTODIAN_CODE_DEFAULT = "CMSHK"

DATE_PATTERNS = [
    "%Y%m%d",
    "%Y-%m-%d",
    "%d-%b-%Y %H:%M:%S",
    "%d-%b-%Y",
    "%m/%d/%Y",
]

ASOF_RE = re.compile(r"SpringGate-TRADE-(\d{8})\.xlsx$")


def get_sheets_service(service_account_json: str):
    credentials_info = json.loads(service_account_json)
    credentials = service_account.Credentials.from_service_account_info(
        credentials_info,
        scopes=["https://www.googleapis.com/auth/spreadsheets.readonly"],
    )
    return build("sheets", "v4", credentials=credentials, cache_discovery=False)


def get_gcs_client(service_account_json: str):
    credentials_info = json.loads(service_account_json)
    credentials = service_account.Credentials.from_service_account_info(
        credentials_info,
        scopes=["https://www.googleapis.com/auth/devstorage.read_write"],
    )
    return storage.Client(project=credentials_info.get("project_id"), credentials=credentials)


def get_values(sheets_service, spreadsheet_id: str, cell_range: str) -> List[List[str]]:
    result = sheets_service.spreadsheets().values().get(
        spreadsheetId=spreadsheet_id,
        range=cell_range,
    ).execute()
    return result.get("values", [])


def parse_date(value: str) -> datetime.date:
    value = (value or "").strip()
    if not value:
        raise ValueError("Empty date value")

    for pattern in DATE_PATTERNS:
        try:
            return datetime.strptime(value, pattern).date()
        except ValueError:
            continue

    raise ValueError(f"Unsupported date format: {value}")


def choose_existing(*candidates: Path) -> Path:
    for path in candidates:
        if path.exists():
            return path
    raise FileNotFoundError(f"None of the candidate paths exists: {[str(p) for p in candidates]}")



def list_bucket_trade_files(gcs_client: storage.Client, bucket_name: str, prefix: str) -> List[str]:
    blobs = gcs_client.list_blobs(bucket_name, prefix=prefix)
    names: List[str] = []
    for blob in blobs:
        if blob.name.endswith(".xlsx"):
            names.append(Path(blob.name).name)
    return names


def detect_latest_trade_date(file_names: Sequence[str]) -> Optional[datetime.date]:
    latest: Optional[datetime.date] = None
    for file_name in file_names:
        match = ASOF_RE.match(file_name)
        if not match:
            continue
        date_value = datetime.strptime(match.group(1), "%Y%m%d").date()
        if latest is None or date_value > latest:
            latest = date_value
    return latest


def upload_trade_file(gcs_client: storage.Client, bucket_name: str, prefix: str, local_path: Path, filename: str) -> None:
    object_name = f"{prefix}{filename}" if prefix else filename
    bucket = gcs_client.bucket(bucket_name)
    blob = bucket.blob(object_name)
    blob.upload_from_filename(
        str(local_path),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    print(f"Uploaded to GCS: gs://{bucket_name}/{object_name}")


def get_raw_trades_rows(
    sheets_service,
    spreadsheet_id: str,
) -> Tuple[List[str], List[Dict[str, str]], str]:
    for tab in ("Raw_Trades", "Raw Trades"):
        rows = get_values(sheets_service, spreadsheet_id, f"{tab}!A:AZ")
        if not rows:
            continue
        header = [h.strip() for h in rows[0]]
        body = rows[1:]
        dict_rows = []
        for row in body:
            if not any(cell.strip() for cell in row):
                continue
            record = {header[i]: row[i] if i < len(row) else "" for i in range(len(header))}
            dict_rows.append(record)
        return header, dict_rows, tab
    raise RuntimeError("No rows found in either 'Raw_Trades' or 'Raw Trades' tab")


def format_investment_code(exchange_code: str, product_code: str) -> str:
    exchange_code = (exchange_code or "").strip().upper()
    product_code = (product_code or "").strip()

    if not product_code:
        return ""

    if exchange_code == "HKEX":
        normalized_code = product_code.upper().replace(".HK", "").replace(" HK", "")
        if normalized_code.isdigit():
            if len(normalized_code) == 5 and normalized_code.startswith("0"):
                normalized_code = normalized_code[1:]
            normalized_code = normalized_code.zfill(4)
        return f"{normalized_code} HK"

    if exchange_code in {"SZMK", "SHMK", "SSE", "SZSE", "MAMK"}:
        return f"{product_code} CH"

    if exchange_code in {"NYSE", "NASDAQ", "AMEX", "USA"}:
        return f"{product_code} US"

    if exchange_code in {"KRX", "KOSPI", "KOSDAQ", "KRW"}:
        normalized_code = product_code.upper().replace(".KS", "").replace(" KS", "")
        return f"{normalized_code} KS"

    if exchange_code in {"TSE", "TOSE", "JPX", "JPY"}:
        return f"{product_code} JP"

    return product_code


def to_float(raw: str) -> float:
    raw = (raw or "").strip().replace(",", "")
    if raw == "":
        return 0.0
    return float(raw)


def write_trade_file(template_path: Path, output_path: Path, rows: Sequence[Dict[str, str]], trade_date: datetime.date) -> None:
    if not output_path.exists():
        shutil.copyfile(template_path, output_path)

    wb = load_workbook(output_path)
    ws = wb.active

    for idx, row in enumerate(rows, start=2):
        t_date = parse_date(row.get("trade_date", ""))
        s_date = parse_date(row.get("settle_date", ""))
        exchange_code = row.get("exchange_code", "")
        product_code = (row.get("product_code") or "").strip()
        investment_code = format_investment_code(exchange_code, product_code)

        bs_type = (row.get("bs_type") or "").strip().upper()
        if bs_type == "B":
            trade_type = "Buy"
        elif bs_type == "S":
            trade_type = "Sell"
        else:
            raise ValueError(f"Unsupported bs_type: {bs_type}")

        seq = f"{idx - 1:03d}"
        external_ref = f"{trade_date:%Y%m%d}_{seq}"

        other_expenses = (
            to_float(row.get("stamp_fee", ""))
            + to_float(row.get("trade_fee", ""))
            + to_float(row.get("trade_levy", ""))
            + to_float(row.get("frc_levy", ""))
            + to_float(row.get("clearing_fee", ""))
        )

        values = {
            "A": PORTFOLIO_DEFAULT,
            "B": FUND_DEFAULT,
            "C": t_date,
            "D": s_date,
            "E": s_date,
            "F": trade_type,
            "G": TRADE_STATUS_DEFAULT,
            "H": investment_code,
            "K": investment_code,
            "N": external_ref,
            "O": (row.get("product_shortname") or "").strip(),
            "P": INVESTMENT_TYPE_DEFAULT,
            "R": (row.get("trade_ccy") or "").strip(),
            "S": (row.get("trade_ccy") or "").strip(),
            "T": (row.get("trade_ccy") or "").strip(),
            "V": to_float(row.get("total_qty", "")),
            "X": to_float(row.get("avg_price", "")),
            "Y": to_float(row.get("gross_amount", "")),
            "Z": to_float(row.get("commission", "")),
            "AA": other_expenses,
            "AB": to_float(row.get("net_amount", "")),
            "AD": BROKER_CODE_DEFAULT,
            "AE": CUSTODIAN_CODE_DEFAULT,
        }

        for col, value in values.items():
            ws[f"{col}{idx}"] = value

        ws[f"A{idx}"].number_format = "@"
        ws[f"B{idx}"].number_format = "@"
        ws[f"C{idx}"].number_format = "yyyy-mm-dd"
        ws[f"D{idx}"].number_format = "yyyy-mm-dd"
        ws[f"E{idx}"].number_format = "yyyy-mm-dd"

    wb.save(output_path)


def main() -> None:
    spreadsheet_id = os.environ["GSHEETS_SPREADSHEET_ID"]
    service_account_json = os.environ["GSHEETS_SERVICE_ACCOUNT_JSON"]
    bucket_name = os.environ["GCS_TRADE_FILES_BUCKET"]
    bucket_prefix = os.environ.get("GCS_TRADE_FILES_PREFIX", "")
    if bucket_prefix and not bucket_prefix.endswith("/"):
        bucket_prefix = f"{bucket_prefix}/"

    repo_root = Path(__file__).resolve().parent.parent
    template_path = choose_existing(repo_root / "main" / "Blank Template.xlsx", repo_root / "Blank Template.xlsx")

    sheets = get_sheets_service(service_account_json)
    gcs_client = get_gcs_client(service_account_json)
    header, raw_rows, source_tab = get_raw_trades_rows(sheets, spreadsheet_id)
    required = [
        "trade_date",
        "settle_date",
        "bs_type",
        "exchange_code",
        "product_code",
        "product_shortname",
        "trade_ccy",
        "total_qty",
        "avg_price",
        "gross_amount",
        "commission",
        "stamp_fee",
        "trade_fee",
        "trade_levy",
        "frc_levy",
        "clearing_fee",
        "net_amount",
    ]
    missing = [c for c in required if c not in header]
    if missing:
        raise RuntimeError(f"Missing required columns in {source_tab}: {missing}")

    grouped: Dict[datetime.date, List[Dict[str, str]]] = {}
    for row in raw_rows:
        t_date = parse_date(row.get("trade_date", ""))
        grouped.setdefault(t_date, []).append(row)

    if not grouped:
        print("No raw trade rows found. Nothing to generate.")
        return

    existing_file_names = list_bucket_trade_files(gcs_client, bucket_name, bucket_prefix)
    local_latest = detect_latest_trade_date(existing_file_names)
    remote_latest = max(grouped.keys())

    if local_latest is None:
        start_date = min(grouped.keys())
    else:
        start_date = local_latest + timedelta(days=1)

    target_dates = sorted(d for d in grouped.keys() if start_date <= d <= remote_latest)
    if not target_dates:
        print(f"No new dates to generate. local_latest={local_latest}, remote_latest={remote_latest}")
        return

    print(f"Generating trade files for dates: {[d.isoformat() for d in target_dates]}")
    print(f"GCS location currently has {len(existing_file_names)} xlsx files.")

    with tempfile.TemporaryDirectory(prefix="trade-files-") as temp_dir:
        temp_path = Path(temp_dir)
        for t_date in target_dates:
            records = grouped[t_date]
            records_sorted = sorted(
                records,
                key=lambda r: (
                    parse_date(r.get("trade_date", "")).isoformat(),
                    (r.get("ref_no") or "").strip(),
                    int((r.get("row_no") or "0").strip() or 0),
                ),
            )
            output_name = f"SpringGate-TRADE-{t_date:%Y%m%d}.xlsx"
            output_path = temp_path / output_name
            write_trade_file(template_path, output_path, records_sorted, t_date)
            upload_trade_file(gcs_client, bucket_name, bucket_prefix, output_path, output_name)
            existing_file_names = [name for name in existing_file_names if name != output_name]
            existing_file_names.append(output_name)


if __name__ == "__main__":
    main()
